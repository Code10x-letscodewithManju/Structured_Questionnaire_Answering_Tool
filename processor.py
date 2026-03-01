import openpyxl
import pandas as pd
from langchain_core.prompts import ChatPromptTemplate

def process_questionnaire(excel_file, retriever, llm):
    """
    Loops through Excel rows, queries the RAG engine, 
    and returns a Workbook and a summary for the UI Review Grid 
    including Evidence Snippets.
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    results_summary = []

    # 1. Map Columns Dynamically
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    q_idx, ans_idx, cite_idx = 1, 2, 3 # Default positions

    for i, header in enumerate(headers):
        h_lower = header.lower()
        if "question" in h_lower: q_idx = i + 1
        if "answer" in h_lower or "response" in h_lower: ans_idx = i + 1
        if "citation" in h_lower or "source" in h_lower: cite_idx = i + 1

    # 2. Grounding prompt
    prompt_template = ChatPromptTemplate.from_template("""
    You are a professional Security Compliance Officer.
    Answer the following question using ONLY the provided context.
    If the context does not contain the answer, strictly state: "Information not found in reference documents."
    
    Context:
    {context}
    
    Question: {question}
    
    RESPONSE FORMAT:
    Provide the answer, then a vertical bar '|', then the source and page.
    Example: The policy requires 256-bit encryption. | Policy_Doc.pdf, Page 4
    """)

    # 3. Iterate through rows
    for row in range(2, ws.max_row + 1):
        question = ws.cell(row=row, column=q_idx).value
        
        if not question or str(question).strip() == "":
            continue
        
        try:
            # Retrieve relevant chunks
            docs = retriever.invoke(str(question))
            
            # FEATURE 2: Extract top 2 snippets as raw evidence
            # This captures the actual text before the AI summarizes it
            evidence_list = []
            for d in docs[:2]:
                snippet = d.page_content[:300].replace('\n', ' ') + "..."
                evidence_list.append(f"[{d.metadata.get('source', 'Doc')}]: {snippet}")
            evidence_text = "\n\n".join(evidence_list)

            # Format context for LLM
            context_str = "\n".join([f"[{d.metadata.get('source')} pg {d.metadata.get('page')}]: {d.page_content}" for d in docs])
            
            # Generate answer
            response = llm.invoke(prompt_template.format(context=context_str, question=question))
            content = response.content
            
            if "|" in content:
                answer, citation = content.split("|", 1)
            else:
                answer, citation = content, "No specific citation found."
            
            # 4. Write back to Excel
            ws.cell(row=row, column=ans_idx).value = answer.strip()
            ws.cell(row=row, column=cite_idx).value = citation.strip()
            
            # 5. Add to summary with Evidence
            results_summary.append({
                "ID": ws.cell(row=row, column=1).value if ws.max_column >= 1 else row-1,
                "Question": question, 
                "Answer": answer.strip(), 
                "Citation": citation.strip(),
                "Evidence Proof": evidence_text  # Added Feature 2
            })
            
        except Exception as e:
            results_summary.append({
                "Question": question, 
                "Answer": "ERROR: Processing Failed", 
                "Citation": "N/A",
                "Evidence Proof": str(e)
            })
        
    return wb, results_summary